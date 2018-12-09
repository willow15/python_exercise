import re

from bs4 import BeautifulSoup
import openpyxl


def get_content(tag):
    tag = str(tag).replace('\n', '')
    match1 = re.match('.+>([\/\:\.\(\)\w\s-]+)</.+', tag)
    if match1:
        return match1.group(1)
    match2 = re.match('.+>([\:\.\(\)\w\s-]+)<.+', tag)
    if match2:
        return match2.group(1)
    return None


def transfer_html_to_xlsx(source_path, destination_path):
    book = openpyxl.Workbook()
    sheet = book.active
    nrow = 0

    with open(source_path) as fp:
        soup = BeautifulSoup(fp, 'lxml')
        tables = soup.body.find_all('table', recursive=False)
        for table in tables:
            trs = table.find_all('tr', recursive=False)
            for tr in trs:
                nrow += 1
                ncol = 0
                tds = tr.find_all('td', recursive=False)
                for td in tds:
                    ncol += 1
                    subtable = td.find('table')
                    if subtable:
                        subtrs = subtable.find_all('tr', recursive=False)
                        for subtr in subtrs:
                            nrow += 1
                            ncol = 0
                            subtds = subtr.find_all('td', recursive=False)
                            for subtd in subtds:
                                ncol += 1
                                sheet.cell(row=nrow, column=ncol).value = get_content(subtd)
                    else:
                        sheet.cell(row=nrow, column=ncol).value = get_content(td)

    book.save(destination_path)


if __name__ == '__main__':
    file_path = 'D:/icbc_report/2018-12-06/Outstanding CA Report_20181205.xls'
    destination_path = 'D:/icbc_report/test.xlsx'
    transfer_html_to_xlsx(file_path, destination_path)
